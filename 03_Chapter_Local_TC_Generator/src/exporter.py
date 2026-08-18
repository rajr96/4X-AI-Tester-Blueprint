"""Export generated test cases as CSV and Excel workbooks."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from datetime import datetime


RESULTS_DIR = Path(__file__).resolve().parent / "Results"
EXPORT_COLUMNS = [
    "test_case_id",
    "scenario",
    "title",
    "priority",
    "type",
    "preconditions",
    "test_data",
    "steps",
    "expected_result",
    "overall_expected_result",
    "assumptions_coverage_gaps",
]


def _value(block: str, labels: tuple[str, ...]) -> str:
    pattern = r"(?:\*\*)?(?:" + "|".join(re.escape(label) for label in labels) + r")(?:\*\*)?\s*:?\s*(.*)"
    match = re.search(pattern, block, re.IGNORECASE)
    return match.group(1).strip(" *") if match else ""


def parse_test_cases(response: str) -> list[dict[str, str]]:
    try:
        payload = json.loads(response)
    except json.JSONDecodeError as exc:
        raise ValueError("Generated response is not valid JSON.") from exc
    rows: list[dict[str, str]] = []
    if isinstance(payload, dict) and isinstance(payload.get("scenario_coverage"), list):
        for category in payload["scenario_coverage"]:
            for case in category.get("test_cases", []):
                rows.append({
                    "test_case_id": str(case["id"]),
                    "scenario": str(category["scenario"]),
                    "title": str(case["title"]),
                    "priority": str(case.get("priority", "")),
                    "type": str(case.get("type", "")),
                    "preconditions": str(case.get("preconditions", "")),
                    "test_data": str(case.get("test_data", "")),
                    "steps": "\n".join(str(step) for step in case.get("steps", [])),
                    "expected_result": str(case["expected_result"]),
                    "overall_expected_result": str(case.get("overall_expected_result", "")),
                    "assumptions_coverage_gaps": str(case.get("assumptions_coverage_gaps", "")),
                })
        return rows
    raise ValueError("Generated response does not match the required JSON structure.")


def _parse_legacy_markdown(markdown: str) -> list[dict[str, str]]:
    blocks = re.split(r"(?=^\*{0,2}\d+[.)]\s+)", markdown, flags=re.MULTILINE)
    rows: list[dict[str, str]] = []
    for block in blocks:
        if not re.search(r"^\*{0,2}\d+[.)]\s+", block):
            continue
        row = {
            "test_case_id": _value(block, ("ID", "Test Case ID")),
            "scenario": "Uncategorized",
            "title": _value(block, ("Title",)),
            "priority": _value(block, ("Priority",)),
            "type": _value(block, ("Type",)),
            "preconditions": _value(block, ("Preconditions", "Pre Conditions", "Precondition")),
            "test_data": _value(block, ("Test Data",)),
            "steps": _value(block, ("Steps",)),
            "expected_result": _value(block, ("Expected Result",)),
            "overall_expected_result": _value(block, ("Overall Expected Result", "Overall Expected result")),
            "assumptions_coverage_gaps": _value(block, ("Assumptions/Coverage Gaps", "Assumptions/Coverage Gap", "Assumptions/ Coverage Gaps")),
        }
        if not row["test_case_id"]:
            heading = re.search(r"(?:Test Case ID|ID)\s*[:#-]?\s*([A-Z0-9_-]+)", block, re.IGNORECASE)
            row["test_case_id"] = heading.group(1) if heading else f"TC-{len(rows) + 1:03d}"
        rows.append(row)
    return rows


def export_test_cases(issue_key: str, response: str) -> tuple[Path, Path]:
    rows = parse_test_cases(response)
    if not 2 <= len(rows) <= 70:
        raise ValueError(f"Expected 2 to 70 exported test cases, got {len(rows)}.")
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{issue_key}_test_cases_{timestamp}"
    csv_path = RESULTS_DIR / f"{filename}.csv"
    xlsx_path = RESULTS_DIR / f"{filename}.xlsx"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"
    sheet.append(EXPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
    for row in rows:
        sheet.append([row[column] for column in EXPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 45)
    workbook.save(xlsx_path)
    return csv_path, xlsx_path